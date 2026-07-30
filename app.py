import psycopg2
import psycopg2.extras
from psycopg2 import IntegrityError
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, Response, session
from datetime import datetime
from functools import wraps
import calendar
import csv
from io import StringIO, BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

app = Flask(__name__)
# Clave secreta necesaria para los mensajes de éxito/error (flash)
app.secret_key = 'mi_clave_secreta_kardex'

# Decorador para proteger rutas/endpoints que solo puede usar un administrador
# ya autenticado (misma sesión que usa la pantalla /admin).
def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return jsonify({'success': False, 'error': 'No autorizado. Debes iniciar sesión como administrador.'}), 403
        return f(*args, **kwargs)
    return decorated

def get_db_connection():
    conn = psycopg2.connect(
        dbname="kardex_jd",
        user="postgres",
        password="fc17181931",
        host="localhost"
    )
    return conn


# 1. Función de validación de IP (Colócala antes de las rutas)
def es_ip_autorizada():
    ip_cliente = request.remote_addr
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT 1 FROM ips_autorizadas WHERE ip_direccion = %s', (ip_cliente,))
    autorizada = cursor.fetchone() is not None
    cursor.close()
    conn.close()
    return autorizada

# 2. Nueva ruta raíz con filtro de seguridad
@app.route('/')
def index():
    if es_ip_autorizada():
        return renderizar_kardex_completo()
    else:
        return redirect(url_for('consultor'))

# 3. La función que contiene toda tu lógica original del index
def renderizar_kardex_completo():
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # Obtener el mes desde la URL
    mes_filtro = request.args.get('mes')
    if not mes_filtro:
        mes_filtro = datetime.now().strftime('%Y-%m')
        
    cursor.execute('SELECT * FROM materiales ORDER BY nombre ASC')
    materiales_db = cursor.fetchall()
    materiales_kardex = []
    
    alertas_rojas = []
    alertas_amarillas = []
    
    totales = {
        'ini_cant': 0, 'ini_total': 0,
        'ing_cant': 0, 'ing_total': 0,
        'sal_cant': 0, 'sal_total': 0,
        'fin_cant': 0, 'fin_total': 0
    }

    hoy = datetime.now()
    try:
        _, ultimo_dia = calendar.monthrange(hoy.year, hoy.month)
        es_fin_de_mes = (ultimo_dia - hoy.day) <= 3
    except Exception:
        es_fin_de_mes = False
        
    for mat in materiales_db:
        mat_id = mat['id']
        cant_saldo = mat['cantidad_inicial']
        precio_promedio = mat['precio_unitario']
        total_saldo = cant_saldo * precio_promedio
        
        cursor.execute('SELECT * FROM movimientos WHERE material_id = %s ORDER BY fecha ASC, id ASC', (mat_id,))
        movimientos = cursor.fetchall()
        
        if mes_filtro != 'todos':
            movs_anteriores = [m for m in movimientos if str(m['fecha']) < f"{mes_filtro}-01"]
            movs_actuales = [m for m in movimientos if str(m['fecha']).startswith(mes_filtro)]
        else:
            movs_anteriores = []
            movs_actuales = movimientos
            
        for mov in movs_anteriores:
            if mov['tipo'] == 'entrada':
                costo_movimiento = mov['cantidad'] * mov['precio_unitario']
                cant_saldo += mov['cantidad']
                total_saldo += costo_movimiento
                if cant_saldo > 0: precio_promedio = total_saldo / cant_saldo
            elif mov['tipo'] == 'salida':
                costo_movimiento = mov['cantidad'] * precio_promedio
                cant_saldo -= mov['cantidad']
                total_saldo -= costo_movimiento
        
        ini_cant, ini_costo, ini_total = cant_saldo, precio_promedio, total_saldo
        
        acum_ingreso_cant, acum_ingreso_total = 0, 0
        acum_salida_cant, acum_salida_total = 0, 0
        
        for mov in movs_actuales:
            if mov['tipo'] == 'entrada':
                costo_movimiento = mov['cantidad'] * mov['precio_unitario']
                cant_saldo += mov['cantidad']
                total_saldo += costo_movimiento
                acum_ingreso_cant += mov['cantidad']
                acum_ingreso_total += costo_movimiento
                if cant_saldo > 0: precio_promedio = total_saldo / cant_saldo
            elif mov['tipo'] == 'salida':
                costo_movimiento = mov['cantidad'] * precio_promedio
                cant_saldo -= mov['cantidad']
                total_saldo -= costo_movimiento
                acum_salida_cant += mov['cantidad']
                acum_salida_total += costo_movimiento

        avg_ingreso = acum_ingreso_total / acum_ingreso_cant if acum_ingreso_cant > 0 else 0
        avg_salida = acum_salida_total / acum_salida_cant if acum_salida_cant > 0 else 0

        if cant_saldo < 2:
            alertas_rojas.append({'nombre': mat['nombre'], 'stock': cant_saldo})
        elif cant_saldo < 5:
            alertas_amarillas.append({'nombre': mat['nombre'], 'stock': cant_saldo})

        materiales_kardex.append({
            'id': mat['id'],
            'nombre': mat['nombre'],
            'descripcion': dict(mat).get('descripcion', ''),
            'drive_link': mat['drive_link'],
            'costo_link': dict(mat).get('costo_link', ''),
            'tipo_material': mat['tipo_material'],
            'unidad': mat['unidad'],
            'ini_cant': ini_cant, 'ini_costo': ini_costo, 'ini_total': ini_total,
            'ing_cant': acum_ingreso_cant, 'ing_costo': avg_ingreso, 'ing_total': acum_ingreso_total,
            'sal_cant': acum_salida_cant, 'sal_costo': avg_salida, 'sal_total': acum_salida_total,
            'fin_cant': cant_saldo, 'fin_costo': precio_promedio, 'fin_total': total_saldo
        })
        
        totales['ini_cant'] += ini_cant; totales['ini_total'] += ini_total
        totales['ing_cant'] += acum_ingreso_cant; totales['ing_total'] += acum_ingreso_total
        totales['sal_cant'] += acum_salida_cant; totales['sal_total'] += acum_salida_total
        totales['fin_cant'] += cant_saldo; totales['fin_total'] += total_saldo

    cursor.execute('SELECT * FROM grupos ORDER BY nombre ASC')
    grupos = cursor.fetchall()
    cursor.close()
    conn.close()
    
    return render_template('index.html', materiales=materiales_kardex, grupos=grupos, mes_filtro=mes_filtro, 
                           alertas_rojas=alertas_rojas, alertas_amarillas=alertas_amarillas, 
                           es_fin_de_mes=es_fin_de_mes, totales=totales)

@app.route('/inventario', methods=['GET', 'POST'])
def inventario():
    if request.method == 'POST':
        nombre = request.form['nombre']
        descripcion = request.form.get('descripcion', '')
        tipo_material = request.form['tipo_material']
        numero_metrico = request.form['numero_metrico']
        origen = request.form['origen']
        empresa = request.form['empresa']
        presentacion = request.form['presentacion']
        unidad = request.form['unidad']
        cantidad_inicial = float(request.form['cantidad_inicial'])
        precio_unitario = float(request.form['precio_unitario'])
        fuente = request.form.get('fuente', '')
        drive_link = request.form.get('drive_link', '')

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO materiales (nombre, descripcion, tipo_material, numero_metrico, origen, empresa, presentacion, unidad, cantidad_inicial, precio_unitario, fuente, drive_link)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (nombre, descripcion, tipo_material, numero_metrico, origen, empresa, presentacion, unidad, cantidad_inicial, precio_unitario, fuente, drive_link))
        conn.commit()
        cursor.close()
        conn.close()

        flash("Éxito: Material agregado correctamente al Inventario.", "success")
        return redirect(url_for('inventario'))

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    # 1. Obtener todos los materiales
    cursor.execute('SELECT * FROM materiales ORDER BY nombre ASC')
    materiales_raw = cursor.fetchall()
    
    # 2. Calcular stock y costo promedio actual para cada uno
    materiales = []
    for mat in materiales_raw:
        m = dict(mat)
        
        # Consultamos el acumulado de movimientos
        cursor.execute('''
            SELECT 
                SUM(CASE WHEN tipo='entrada' THEN cantidad ELSE -cantidad END) as mov_cant,
                SUM(CASE WHEN tipo='entrada' THEN (cantidad * precio_unitario) ELSE 0 END) as total_entradas,
                SUM(CASE WHEN tipo='entrada' THEN cantidad ELSE 0 END) as cant_entradas
            FROM movimientos WHERE material_id = %s
        ''', (m['id'],))
        res = cursor.fetchone()
        
        # Cálculo: Stock = Inicial + Entradas - Salidas
        m['stock_actual'] = m['cantidad_inicial'] + (res['mov_cant'] or 0)
        
        # Cálculo: Costo Promedio (Total Invertido / Cantidad Total)
        total_acumulado = (m['cantidad_inicial'] * m['precio_unitario']) + (res['total_entradas'] or 0)
        total_cantidad = m['cantidad_inicial'] + (res['cant_entradas'] or 0)
        
        m['costo_promedio_actual'] = (total_acumulado / total_cantidad) if total_cantidad > 0 else m['precio_unitario']
        
        materiales.append(m)

    cursor.execute('SELECT * FROM grupos ORDER BY nombre ASC')
    grupos = cursor.fetchall()
    cursor.execute('SELECT * FROM proveedores ORDER BY nombre ASC')
    proveedores = cursor.fetchall()
    cursor.execute('SELECT * FROM fuentes ORDER BY nombre ASC')
    fuentes = cursor.fetchall()
    
    cursor.close()
    conn.close()
    
    return render_template('inventario.html', materiales=materiales, grupos=grupos, proveedores=proveedores, fuentes=fuentes)

@app.route('/editar_material', methods=['POST'])
def editar_material():
    if request.method == 'POST':
        id_material = int(request.form['id'])
        nombre = request.form['nombre']
        descripcion = request.form.get('descripcion', '')
        tipo_material = request.form['tipo_material']
        numero_metrico = request.form['numero_metrico']
        origen = request.form['origen']
        empresa = request.form['empresa']
        presentacion = request.form['presentacion']
        unidad = request.form['unidad']
        cantidad_inicial = float(request.form['cantidad_inicial'])
        precio_unitario = float(request.form['precio_unitario'])
        fuente = request.form.get('fuente', '')
        drive_link = request.form.get('drive_link', '')

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE materiales 
            SET nombre = %s, descripcion = %s, tipo_material = %s, numero_metrico = %s, origen = %s, empresa = %s, presentacion = %s, unidad = %s, cantidad_inicial = %s, precio_unitario = %s, fuente = %s, drive_link = %s
            WHERE id = %s
        ''', (nombre, descripcion, tipo_material, numero_metrico, origen, empresa, presentacion, unidad, cantidad_inicial, precio_unitario, fuente, drive_link, id_material))
        conn.commit()
        cursor.close()
        conn.close()

        flash("Éxito: Material actualizado correctamente.", "success")
        return redirect(url_for('inventario'))

@app.route('/agregar_grupo_ajax', methods=['POST'])
def agregar_grupo_ajax():
    nombre = request.json.get('nombre')
    if not nombre:
        return jsonify({'success': False, 'error': 'El nombre está vacío'})
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('INSERT INTO grupos (nombre) VALUES (%s) RETURNING id', (nombre,))
        nuevo_id = cursor.fetchone()[0]
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'id': nuevo_id, 'nombre': nombre})
    except IntegrityError:
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'error': 'El grupo ya existe'})

@app.route('/editar_grupo_ajax', methods=['POST'])
def editar_grupo_ajax():
    data = request.json
    id = data.get('id')
    nombre = data.get('nombre')
    nombre_viejo = data.get('nombre_viejo')
    
    if not nombre:
        return jsonify({'success': False, 'error': 'El nombre está vacío'})
        
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Actualizar grupo
        cursor.execute('UPDATE grupos SET nombre = %s WHERE id = %s', (nombre, id))
        # Actualizar todos los materiales que usaban este grupo al nuevo nombre
        if nombre != nombre_viejo:
            cursor.execute('UPDATE materiales SET tipo_material = %s WHERE tipo_material = %s', (nombre, nombre_viejo))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except IntegrityError:
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'error': 'El grupo ya existe'})
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/eliminar_grupo_ajax', methods=['POST'])
def eliminar_grupo_ajax():
    data = request.json
    id = data.get('id')
    pin = data.get('pin')
    
    if pin != '1234':
        return jsonify({'success': False, 'error': 'PIN incorrecto'})
        
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('DELETE FROM grupos WHERE id = %s', (id,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/actualizar_vinculo_ajax', methods=['POST'])
def actualizar_vinculo_ajax():
    data = request.json
    material_id = data.get('material_id')
    link = data.get('link', '')
    tipo = data.get('tipo', 'nombre') # 'nombre' o 'costo'

    if not material_id:
        return jsonify({'success': False, 'error': 'ID de material no proporcionado'})

    columna = 'drive_link' if tipo == 'nombre' else 'costo_link'

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute(f'UPDATE materiales SET {columna} = %s WHERE id = %s', (link, material_id))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/agregar_proveedor_ajax', methods=['POST'])
def agregar_proveedor_ajax():
    nit = request.json.get('nit', '')
    nombre = request.json.get('nombre')
    if not nombre:
        return jsonify({'success': False, 'error': 'El nombre está vacío'})
    
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('INSERT INTO proveedores (nit, nombre) VALUES (%s, %s) RETURNING id', (nit, nombre))
        nuevo_id = cursor.fetchone()[0]
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True, 'id': nuevo_id, 'nombre': nombre})
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/editar_proveedor_ajax', methods=['POST'])
def editar_proveedor_ajax():
    data = request.json
    id = data.get('id')
    nit = data.get('nit', '')
    nombre = data.get('nombre')
    nombre_viejo = data.get('nombre_viejo')
    
    if not nombre:
        return jsonify({'success': False, 'error': 'El nombre está vacío'})
        
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # Actualizar proveedor
        cursor.execute('UPDATE proveedores SET nit = %s, nombre = %s WHERE id = %s', (nit, nombre, id))
        # Actualizar todos los materiales que usaban este proveedor al nuevo nombre
        if nombre != nombre_viejo:
            cursor.execute('UPDATE materiales SET empresa = %s WHERE empresa = %s', (nombre, nombre_viejo))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/eliminar_proveedor_ajax', methods=['POST'])
def eliminar_proveedor_ajax():
    data = request.json
    id = data.get('id')
    pin = data.get('pin')
    
    if pin != '1234':
        return jsonify({'success': False, 'error': 'PIN incorrecto'})
        
    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('DELETE FROM proveedores WHERE id = %s', (id,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/agregar_entrada', methods=['POST'])
def agregar_entrada():
    if request.method == 'POST':
        material_id = int(request.form['material_id'])
        cantidad = float(request.form['cantidad'])
        
        # Asigna 0.0 temporalmente si el precio viene vacío
        precio_str = request.form.get('precio', '').strip()
        precio = float(precio_str) if precio_str else 0.0
        
        fecha = request.form.get('fecha')
        fecha_factura = request.form.get('fecha_factura', '')
        
        tipo_documento = request.form.get('tipo_documento') 
        # En tu form, este input captura lo que el usuario escribe (ya sea Factura u Orden)
        numero_documento = request.form.get('numero_documento', '').strip() 

        if not fecha:
            fecha = datetime.now().strftime('%Y-%m-%d')

        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
        if tipo_documento == 'devolucion':
            # Buscamos la salida donde el 'documento' sea igual a la Orden ingresada
            cursor.execute('''
                SELECT precio_unitario FROM movimientos
                WHERE material_id = %s AND tipo = 'salida' AND documento = %s
                ORDER BY id DESC
                LIMIT 1
            ''', (material_id, numero_documento))
            salida = cursor.fetchone()
            
            if salida:
                precio = salida['precio_unitario'] # Reemplazamos por el costo exacto
            else:
                cursor.close()
                conn.close()
                flash(f"Error: No se encontró una salida asociada a la Orden '{numero_documento}'.", "error")
                return redirect(url_for('index'))
            
            documento_bd = "Devolución"
        else:
            documento_bd = "Factura"

        # Registrar la entrada
        cursor.execute('''
            INSERT INTO movimientos (material_id, tipo, cantidad, precio_unitario, fecha, documento, numero_documento, fecha_factura)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ''', (material_id, 'entrada', cantidad, precio, fecha, documento_bd, numero_documento, fecha_factura))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash(f"Éxito: {documento_bd} registrada correctamente.", "success")
        if request.form.get('origen') == 'vista_entradas':
            return redirect(url_for('entradas'))
        return redirect(url_for('index'))

# --- NUEVA RUTA PARA QUE EL MODAL SEA INTELIGENTE ---
@app.route('/api/precio_devolucion')
def api_precio_devolucion():
    material_id = request.args.get('material_id', type=int)
    orden = request.args.get('orden', '')

    if not material_id or not orden:
        return jsonify({'success': False, 'error': 'Faltan datos'})

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    # Busca el precio original basado en la orden (documento)
    cursor.execute('''
        SELECT precio_unitario FROM movimientos
        WHERE material_id = %s AND tipo = 'salida' AND documento = %s
        ORDER BY id DESC
        LIMIT 1
    ''', (material_id, orden))
    salida = cursor.fetchone()
    cursor.close()
    conn.close()

    if salida:
        return jsonify({'success': True, 'precio': salida['precio_unitario']})
    else:
        return jsonify({'success': False, 'error': 'No encontrada'})

@app.route('/agregar_salida', methods=['POST'])
def agregar_salida():
    if request.method == 'POST':
        material_id = int(request.form['material_id'])
        cantidad_a_sacar = float(request.form['cantidad'])
        fecha = request.form.get('fecha')
        # Ahora el documento es fijo "Orden"
        documento = request.form.get('documento', 'Orden')
        # El número de documento es el correlativo único
        numero_documento = request.form.get('numero_documento', '').strip()
        departamento = request.form.get('departamento', '')
        solicitante = request.form.get('solicitante', '')

        if not fecha:
            fecha = datetime.now().strftime('%Y-%m-%d')

        if not numero_documento:
            flash("Error: El correlativo es obligatorio.", "error")
            return redirect(url_for('index'))

        conn = get_db_connection()
        cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    

        # 2. Validar existencias actuales
        cursor.execute('SELECT * FROM materiales WHERE id = %s', (material_id,))
        material = cursor.fetchone()
        cant_actual = material['cantidad_inicial']
        total_actual = material['cantidad_inicial'] * material['precio_unitario']
        precio_promedio = material['precio_unitario']
        
        cursor.execute('SELECT * FROM movimientos WHERE material_id = %s ORDER BY fecha ASC, id ASC', (material_id,))
        movimientos = cursor.fetchall()
        for mov in movimientos:
            if mov['tipo'] == 'entrada':
                cant_actual += mov['cantidad']
                total_actual += (mov['cantidad'] * mov['precio_unitario'])
                if cant_actual > 0: precio_promedio = total_actual / cant_actual
            elif mov['tipo'] == 'salida':
                cant_actual -= mov['cantidad']
                total_actual -= (mov['cantidad'] * precio_promedio)

        # BLOQUEO SI NO HAY STOCK SUFICIENTE
        if cantidad_a_sacar > cant_actual:
            cursor.close()
            conn.close()
            flash(f"Error: No puedes sacar {cantidad_a_sacar} unidades. Solo hay {cant_actual} disponibles del material '{material['nombre']}'.", "error")
            
            if request.form.get('origen') == 'vista_salidas':
                return redirect(url_for('salidas'))
            return redirect(url_for('index'))

        # Si hay stock y correlativo único, registrar la salida
        cursor.execute('''
            INSERT INTO movimientos (material_id, tipo, cantidad, precio_unitario, fecha, documento, numero_documento, departamento, solicitante)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        ''', (material_id, 'salida', cantidad_a_sacar, precio_promedio, fecha, documento, numero_documento, departamento, solicitante))
        
        conn.commit()
        cursor.close()
        conn.close()
        
        flash(f"Éxito: Salida registrada correctamente con correlativo {numero_documento}.", "success")
        if request.form.get('origen') == 'vista_salidas':
            return redirect(url_for('salidas'))
        return redirect(url_for('index'))
    

@app.route('/eliminar_material/<int:id>', methods=['POST'])
def eliminar_material(id):
    if request.method == 'POST':
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM movimientos WHERE material_id = %s', (id,))
        cursor.execute('DELETE FROM materiales WHERE id = %s', (id,))
        conn.commit()
        cursor.close()
        conn.close()
        flash("Éxito: Material eliminado correctamente.", "success")
        return redirect(url_for('inventario'))

@app.route('/entradas')
def entradas():
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute('SELECT * FROM materiales ORDER BY nombre ASC')
    materiales = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('entradas.html', materiales=materiales)

@app.route('/salidas')
def salidas():
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute('SELECT * FROM materiales ORDER BY nombre ASC')
    materiales = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('salidas.html', materiales=materiales)

@app.route('/reporte')
def reporte():
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    cursor.execute('SELECT * FROM materiales ORDER BY nombre ASC')
    materiales = cursor.fetchall()
    
    selected_material_id = request.args.get('material_id', type=int)
    mes_filtro = request.args.get('mes')
    if not mes_filtro:
        mes_filtro = datetime.now().strftime('%Y-%m')
        
    reporte_datos = None
    
    if selected_material_id:
        cursor.execute('SELECT * FROM materiales WHERE id = %s', (selected_material_id,))
        mat = cursor.fetchone()
        if mat:
            mat_id = mat['id']
            cant_saldo = mat['cantidad_inicial']
            precio_promedio = mat['precio_unitario']
            total_saldo = cant_saldo * precio_promedio
            
            cursor.execute('SELECT * FROM movimientos WHERE material_id = %s ORDER BY fecha ASC, id ASC', (mat_id,))
            movimientos = cursor.fetchall()
            
            if mes_filtro != 'todos':
                movs_anteriores = [m for m in movimientos if str(m['fecha']) < f"{mes_filtro}-01"]
                movs_actuales = [m for m in movimientos if str(m['fecha']).startswith(mes_filtro)]
            else:
                movs_anteriores = []
                movs_actuales = movimientos
                
            for mov in movs_anteriores:
                if mov['tipo'] == 'entrada':
                    costo_movimiento = mov['cantidad'] * mov['precio_unitario']
                    cant_saldo += mov['cantidad']
                    total_saldo += costo_movimiento
                    if cant_saldo > 0: precio_promedio = total_saldo / cant_saldo
                elif mov['tipo'] == 'salida':
                    costo_movimiento = mov['cantidad'] * precio_promedio
                    cant_saldo -= mov['cantidad']
                    total_saldo -= costo_movimiento

            filas_kardex = []
            # Primera fila: El saldo inicial o anterior según el filtro
            titulo_saldo = 'Saldo Inicial' if mes_filtro == 'todos' else f'Saldo Anterior ({mes_filtro})'
            filas_kardex.append({
                'fecha': '-', 'detalle': titulo_saldo,
                'ing_cant': '', 'ing_costo': '', 'ing_total': '',
                'sal_cant': '', 'sal_costo': '', 'sal_total': '',
                'saldo_cant': cant_saldo, 'saldo_costo': precio_promedio, 'saldo_total': total_saldo
            })
            
            for mov in movs_actuales:
                doc_info = ""
                if mov['documento'] and mov['numero_documento']:
                    doc_info = f" ({mov['documento']} #{mov['numero_documento']})"
                elif mov['documento']:
                    doc_info = f" ({mov['documento']})"
                    
                if mov['tipo'] == 'entrada':
                    costo_movimiento = mov['cantidad'] * mov['precio_unitario']
                    cant_saldo += mov['cantidad']
                    total_saldo += costo_movimiento
                    if cant_saldo > 0:
                        precio_promedio = total_saldo / cant_saldo
                    
                    filas_kardex.append({
                        'fecha': mov['fecha'], 'detalle': f"Entrada / Compra{doc_info}",
                        'ing_cant': mov['cantidad'], 'ing_costo': mov['precio_unitario'], 'ing_total': costo_movimiento,
                        'sal_cant': '', 'sal_costo': '', 'sal_total': '',
                        'saldo_cant': cant_saldo, 'saldo_costo': precio_promedio, 'saldo_total': total_saldo
                    })
                elif mov['tipo'] == 'salida':
                    costo_movimiento = mov['cantidad'] * precio_promedio
                    cant_saldo -= mov['cantidad']
                    total_saldo -= costo_movimiento
                    
                    filas_kardex.append({
                        'fecha': mov['fecha'], 'detalle': f"Salida / Egreso{doc_info}",
                        'ing_cant': '', 'ing_costo': '', 'ing_total': '',
                        'sal_cant': mov['cantidad'], 'sal_costo': precio_promedio, 'sal_total': costo_movimiento,
                        'saldo_cant': cant_saldo, 'saldo_costo': precio_promedio, 'saldo_total': total_saldo
                    })
                    
            reporte_datos = {'material': mat, 'filas': filas_kardex}
    
    cursor.close()
    conn.close()
    return render_template('reporte.html', materiales=materiales, reporte_datos=reporte_datos, selected_material_id=selected_material_id, mes_filtro=mes_filtro)

# --- RUTAS DE EXPORTACIÓN A EXCEL (CSV) ---
@app.route('/exportar_inventario')
def exportar_inventario():
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute('SELECT * FROM materiales ORDER BY nombre ASC')
    materiales = cursor.fetchall()
    cursor.close()
    conn.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Estilos basados en la interfaz (CSS)
    fill_hdr_gris = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    font_hdr_gris = Font(color="475569", bold=True)
    alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    alignment_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    border_thin = Border(left=Side(style='thin', color='E2E8F0'), 
                         right=Side(style='thin', color='E2E8F0'), 
                         top=Side(style='thin', color='E2E8F0'), 
                         bottom=Side(style='thin', color='E2E8F0'))

    headers = ['Nombre', 'Descripción', 'Grupo', 'No. Metrico', 'Origen', 'Fuente', 'Proveedor', 'Presentacion', 'Unidad', 'Existencia Inicial', 'Costo Unitario (Q)']
    
    ws.append(headers)
    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = fill_hdr_gris
        cell.font = font_hdr_gris
        # Alinear a la izquierda todas las columnas de texto
        cell.alignment = alignment_left if col_num <= 9 else alignment_right
        cell.border = border_thin

    for idx, mat in enumerate(materiales, 1):
        row = [mat['nombre'], mat['descripcion'], mat['tipo_material'], mat['numero_metrico'], mat['origen'], mat['fuente'], mat['empresa'], mat['presentacion'], mat['unidad'], mat['cantidad_inicial'], round(mat['precio_unitario'], 2)]
        ws.append(row)
        for col_num, cell in enumerate(ws[ws.max_row], 1):
            cell.alignment = alignment_left if col_num <= 9 else alignment_right
            cell.border = border_thin
    # Ajustar ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    
    return Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment; filename=Plantilla_Inventario.xlsx'})

@app.route('/exportar_kardex')
def exportar_kardex():
    mes_filtro = request.args.get('mes')
    if not mes_filtro:
        mes_filtro = datetime.now().strftime('%Y-%m')
    
    # 1. Configurar nombres y libro con manejo de errores
    meses_es = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    
    if mes_filtro == 'todos':
        nombre_archivo = "Kardex General Completo.xlsx"
    else:
        try:
            año, mes_num = mes_filtro.split('-')
            mes_nombre = meses_es[int(mes_num)]
            nombre_archivo = f"Kardex General Mes de {mes_nombre} de {año}.xlsx"
        except:
            nombre_archivo = "Kardex_General.xlsx"
            
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    cursor.execute('SELECT * FROM materiales ORDER BY nombre ASC')
    materiales = cursor.fetchall()
    
    wb = openpyxl.Workbook()
    # ... (El resto de tu código de estilos y generación de Excel sigue exactamente igual hacia abajo)
    
    # --- ESTILOS ---
    fill_verde = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    fill_naranja = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
    fill_azul = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    fill_gris = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    bold_font = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # --- HOJA 1: KARDEX DETALLADO ---
    ws_kardex = wb.active
    ws_kardex.title = "Kardex Detallado"
    
    # Encabezados con colores
    headers = ['Nombre', 'Descripción', 'Grupo', 'Inicial Cant.', 'Inicial Total', 'Entradas Cant.', 'Entradas Total', 'Salidas Cant.', 'Salidas Total', 'Actual Cant.', 'Actual Total']
    ws_kardex.append(headers)
    
    for col_num, cell in enumerate(ws_kardex[1], 1):
        cell.fill = fill_gris
        cell.font = bold_font
        cell.border = border

    # --- HOJA 2: INVENTARIO ACTUAL ---
    ws_inv = wb.create_sheet(title="Inventario Actual")
    headers_inv = ['Nombre', 'Descripción', 'Grupo', 'Stock Actual', 'Costo Prom. Actual', 'Valor Total Actual']
    ws_inv.append(headers_inv)
    for cell in ws_inv[1]:
        cell.fill = fill_gris
        cell.font = bold_font
        cell.border = border

    # --- PROCESAMIENTO ---
    for mat in materiales:
        # Calcular stock y costo actual
        cursor.execute('''SELECT SUM(CASE WHEN tipo='entrada' THEN cantidad ELSE -cantidad END) as mov_cant,
                          AVG(precio_unitario) as costo_prom
                          FROM movimientos WHERE material_id = %s''', (mat['id'],))
        res = cursor.fetchone()
        
        # SOLUCIÓN: Convertir todo explícitamente a float para evitar el choque de tipos
        cant_inicial = float(mat['cantidad_inicial'])
        precio_uni = float(mat['precio_unitario'])
        mov_cant = float(res['mov_cant'] or 0)
        costo_prom = float(res['costo_prom'] or mat['precio_unitario'])
        
        stock_actual = cant_inicial + mov_cant
        
        # Llenar Hoja Inventario
        row_inv = [
            mat['nombre'], 
            mat['descripcion'], 
            mat['tipo_material'], 
            stock_actual, 
            round(costo_prom, 2), 
            round(stock_actual * costo_prom, 2)
        ]
        ws_inv.append(row_inv)
        
        # Llenar Hoja Kardex
        row_kardex = [
            mat['nombre'], 
            mat['descripcion'], 
            mat['tipo_material'], 
            cant_inicial, 
            round(cant_inicial * precio_uni, 2), 
            0, 0, 0, 0, 
            stock_actual, 
            round(stock_actual * costo_prom, 2)
        ]
        ws_kardex.append(row_kardex)

    # --- APLICAR COLORES A COLUMNAS DE LA HOJA KARDEX ---
    for row in ws_kardex.iter_rows(min_row=2):
        for col_num, cell in enumerate(row, 1):
            cell.border = border
            if 6 <= col_num <= 7: cell.fill = fill_verde # Entradas
            if 8 <= col_num <= 9: cell.fill = fill_naranja # Salidas
            if 10 <= col_num <= 11: cell.fill = fill_azul # Actuales

    cursor.close()
    conn.close()
    
    output = BytesIO()
    wb.save(output)
    return Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                    headers={'Content-Disposition': f'attachment; filename="Kardex General Mes de {mes_nombre} de {año}.xlsx"'})
@app.route('/cargar_excel', methods=['GET', 'POST'])
def cargar_excel():
    if request.method == 'POST':
        if 'archivo_excel' not in request.files:
            flash('Error: No se encontró el archivo en la solicitud.', 'error')
            return redirect(request.url)
        
        file = request.files['archivo_excel']
        
        if file.filename == '':
            flash('Error: No se seleccionó ningún archivo.', 'error')
            return redirect(request.url)

        if file and file.filename.endswith('.xlsx'):
            try:
                conn = get_db_connection()
                cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active
                
                rows_processed = 0
                rows_imported = 0
                rows_skipped = 0
                
                sql_insert = '''
                    INSERT INTO materiales (nombre, descripcion, tipo_material, numero_metrico, origen, fuente, empresa, presentacion, unidad, cantidad_inicial, precio_unitario, drive_link)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                '''
                
                for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    rows_processed += 1
                    try:
                        if len(row) < 11:
                            rows_skipped += 1
                            continue

                        (nombre_raw, descripcion, tipo_material_raw, numero_metrico, origen, 
                         fuente_raw, empresa_raw, presentacion, unidad, cantidad_inicial_raw, precio_unitario_raw) = row[:11]

                        # --- VALIDACIÓN DE DATOS OBLIGATORIOS ---
                        if not all([nombre_raw, tipo_material_raw, fuente_raw, empresa_raw, cantidad_inicial_raw is not None, precio_unitario_raw is not None]):
                            rows_skipped += 1
                            continue

                        # --- LIMPIEZA Y CREACIÓN AUTOMÁTICA DE ENTIDADES ---
                        nombre = str(nombre_raw).strip()
                        tipo_material = str(tipo_material_raw).strip()
                        fuente = str(fuente_raw).strip()
                        empresa = str(empresa_raw).strip()

                        # En PostgreSQL, ON CONFLICT (nombre) DO NOTHING requiere que la columna tenga restricción UNIQUE
                        cursor.execute('INSERT INTO grupos (nombre) VALUES (%s) ON CONFLICT (nombre) DO NOTHING', (tipo_material,))
                        cursor.execute('INSERT INTO fuentes (nombre) VALUES (%s) ON CONFLICT (nombre) DO NOTHING', (fuente,))
                        
                        cursor.execute('SELECT id FROM proveedores WHERE nombre = %s', (empresa,))
                        prov_exists = cursor.fetchone()
                        if not prov_exists:
                            cursor.execute('INSERT INTO proveedores (nombre, nit) VALUES (%s, %s)', (empresa, ''))

                        cantidad_inicial = float(cantidad_inicial_raw)
                        precio_unitario = float(precio_unitario_raw)

                        values_to_insert = (nombre, descripcion, tipo_material, numero_metrico, origen, fuente, empresa, presentacion, unidad, cantidad_inicial, precio_unitario, '')
                        cursor.execute(sql_insert, values_to_insert)
                        rows_imported += 1

                    except (ValueError, TypeError):
                        rows_skipped += 1
                        continue
                
                conn.commit()
                cursor.close()
                conn.close()
                
                flash_message = f"Éxito: Carga completada. Se importaron {rows_imported} materiales."
                if rows_skipped > 0:
                    flash_message += f" Se omitieron {rows_skipped} filas por datos faltantes o formato incorrecto."
                flash(flash_message, "success")

            except Exception as e:
                flash(f"Error: Ocurrió un problema al procesar el archivo Excel: {e}", "error")
            return redirect(url_for('inventario'))

    return render_template('carga_masiva.html')

@app.route('/admin', methods=['GET', 'POST'])
def admin():
    # --- SISTEMA DE LOGIN PARA LA PANTALLA DE ADMIN ---
    if not session.get('admin_logged_in'):
        if request.method == 'POST':
            if request.form.get('admin_password') == 'laFabrica1': # <- Contraseña de administrador
                session['admin_logged_in'] = True
                flash("Acceso concedido.", "success")
                return redirect(url_for('admin'))
            elif request.form.get('admin_password'):
                flash("Error: Contraseña incorrecta.", "error")
        return render_template('admin.html', login_required=True)

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    if request.method == 'POST':
        accion = request.form.get('accion')
        
        if accion == 'logout':
            session.pop('admin_logged_in', None)
            flash("Sesión de administrador cerrada.", "success")
            return redirect(url_for('index'))
            
        if accion == 'grupo':
            try:
                cursor.execute('INSERT INTO grupos (nombre) VALUES (%s)', (request.form['nombre_grupo'],))
                flash("Éxito: Grupo agregado correctamente.", "success")
            except IntegrityError:
                conn.rollback()
                flash("Error: El grupo ya existe.", "error")
                
        elif accion == 'proveedor':
            cursor.execute('INSERT INTO proveedores (nit, nombre) VALUES (%s, %s)', 
                           (request.form['nit'], request.form['nombre']))
            flash("Éxito: Proveedor agregado correctamente.", "success")
            
        elif accion == 'fuente':
            try:
                cursor.execute('INSERT INTO fuentes (nombre) VALUES (%s)', (request.form['nombre_fuente'],))
                flash("Éxito: Fuente agregada correctamente.", "success")
            except IntegrityError:
                conn.rollback()
                flash("Error: La fuente ya existe.", "error")

        elif accion == 'agregar_ip':
            cursor.execute('INSERT INTO ips_autorizadas (ip_direccion, descripcion) VALUES (%s, %s)', 
                           (request.form['nueva_ip'], request.form['desc_ip']))
            conn.commit()
            flash("IP agregada a la lista blanca.", "success")

        conn.commit()
        return redirect(url_for('admin'))
        
    cursor.execute('SELECT * FROM grupos ORDER BY nombre ASC')
    grupos = cursor.fetchall()
    
    cursor.execute('SELECT * FROM proveedores ORDER BY nombre ASC')
    proveedores = cursor.fetchall()
    
    cursor.execute('SELECT * FROM fuentes ORDER BY nombre ASC')
    fuentes = cursor.fetchall()

    cursor.execute('SELECT id, nombre FROM materiales ORDER BY nombre ASC')
    materiales = cursor.fetchall()

    cursor.execute('SELECT * FROM ips_autorizadas ORDER BY id DESC')
    ips = cursor.fetchall()

    cursor.close()
    conn.close()
    return render_template('admin.html', grupos=grupos, proveedores=proveedores, fuentes=fuentes, materiales=materiales, ips=ips)


# --- GESTIÓN DE MOVIMIENTOS (ENTRADAS/SALIDAS) DESDE EL ADMIN ---
# Solo accesible con sesión de administrador iniciada (admin_required).

def _movimiento_a_dict(m):
    """Convierte una fila de movimientos (DictRow) a un dict serializable a JSON."""
    d = dict(m)
    d['cantidad'] = float(d['cantidad']) if d.get('cantidad') is not None else 0
    d['precio_unitario'] = float(d['precio_unitario']) if d.get('precio_unitario') is not None else 0
    d['fecha'] = str(d['fecha']) if d.get('fecha') else ''
    d['fecha_factura'] = str(d['fecha_factura']) if d.get('fecha_factura') else ''
    return d

@app.route('/admin/movimientos')
@admin_required
def admin_listar_movimientos():
    material_id = request.args.get('material_id', type=int)
    tipo = request.args.get('tipo')
    mes = request.args.get('mes')

    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    query = '''
        SELECT mov.*, mat.nombre AS material_nombre, mat.unidad AS material_unidad
        FROM movimientos mov
        JOIN materiales mat ON mat.id = mov.material_id
        WHERE 1=1
    '''
    params = []
    if material_id:
        query += ' AND mov.material_id = %s'
        params.append(material_id)
    if tipo in ('entrada', 'salida'):
        query += ' AND mov.tipo = %s'
        params.append(tipo)
    if mes:
        query += " AND to_char(mov.fecha, 'YYYY-MM') = %s"
        params.append(mes)
    query += ' ORDER BY mov.fecha DESC, mov.id DESC LIMIT 300'

    cursor.execute(query, params)
    movimientos = [_movimiento_a_dict(m) for m in cursor.fetchall()]
    cursor.close()
    conn.close()
    return jsonify({'success': True, 'movimientos': movimientos})

@app.route('/admin/movimiento/editar', methods=['POST'])
@admin_required
def admin_editar_movimiento():
    data = request.json or {}
    id_mov = data.get('id')

    if not id_mov:
        return jsonify({'success': False, 'error': 'Movimiento no especificado.'})

    try:
        cantidad = float(data.get('cantidad'))
        precio_unitario = float(data.get('precio_unitario'))
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'Cantidad o precio unitario inválidos.'})

    fecha = data.get('fecha') or None
    if not fecha:
        return jsonify({'success': False, 'error': 'La fecha es obligatoria.'})

    documento = data.get('documento', '')
    numero_documento = (data.get('numero_documento') or '').strip()
    fecha_factura = data.get('fecha_factura') or None
    departamento = data.get('departamento', '')
    solicitante = data.get('solicitante', '')

    if not numero_documento:
        return jsonify({'success': False, 'error': 'El correlativo/documento es obligatorio.'})

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        # El correlativo debe seguir siendo único, excluyendo el propio registro
        cursor.execute('SELECT id FROM movimientos WHERE numero_documento = %s AND id != %s', (numero_documento, id_mov))
        if cursor.fetchone():
            cursor.close()
            conn.close()
            return jsonify({'success': False, 'error': f"El correlativo '{numero_documento}' ya está en uso por otro movimiento."})

        cursor.execute('''
            UPDATE movimientos
            SET cantidad = %s, precio_unitario = %s, fecha = %s, documento = %s,
                numero_documento = %s, fecha_factura = %s, departamento = %s, solicitante = %s
            WHERE id = %s
        ''', (cantidad, precio_unitario, fecha, documento, numero_documento, fecha_factura, departamento, solicitante, id_mov))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'error': str(e)})

@app.route('/admin/movimiento/eliminar', methods=['POST'])
@admin_required
def admin_eliminar_movimiento():
    data = request.json or {}
    id_mov = data.get('id')
    pin = data.get('pin')

    if pin != '1234':
        return jsonify({'success': False, 'error': 'PIN incorrecto.'})
    if not id_mov:
        return jsonify({'success': False, 'error': 'Movimiento no especificado.'})

    conn = get_db_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('DELETE FROM movimientos WHERE id = %s', (id_mov,))
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        cursor.close()
        conn.close()
        return jsonify({'success': False, 'error': str(e)})


@app.route('/eliminar_grupo/<int:id>', methods=['POST'])
def eliminar_grupo(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM grupos WHERE id = %s', (id,))
    conn.commit()
    cursor.close()
    conn.close()
    flash("Éxito: Grupo eliminado correctamente.", "success")
    return redirect(url_for('admin'))

@app.route('/eliminar_proveedor/<int:id>', methods=['POST'])
def eliminar_proveedor(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM proveedores WHERE id = %s', (id,))
    conn.commit()
    cursor.close()
    conn.close()
    flash("Éxito: Proveedor eliminado correctamente.", "success")
    return redirect(url_for('admin'))

@app.route('/eliminar_fuente/<int:id>', methods=['POST'])
def eliminar_fuente(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM fuentes WHERE id = %s', (id,))
    conn.commit()
    cursor.close()
    conn.close()
    flash("Éxito: Fuente eliminada correctamente.", "success")
    return redirect(url_for('admin'))

@app.route('/eliminar_ip/<int:id>', methods=['POST'])
def eliminar_ip(id):
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM ips_autorizadas WHERE id = %s', (id,))
    conn.commit()
    cursor.close()
    conn.close()
    flash("Éxito: IP eliminada de la lista blanca.", "success")
    return redirect(url_for('admin'))

@app.route('/consultor')
def consultor():
    conn = get_db_connection()
    cursor = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)
    
    cursor.execute('SELECT * FROM materiales ORDER BY nombre ASC')
    materiales_db = cursor.fetchall()
    stock_materiales = []

    for mat in materiales_db:
        mat_id = mat['id']
        cant_saldo = mat['cantidad_inicial']
        cursor.execute('SELECT tipo, cantidad FROM movimientos WHERE material_id = %s', (mat_id,))
        movimientos = cursor.fetchall()
        
        for mov in movimientos:
            if mov['tipo'] == 'entrada':
                cant_saldo += mov['cantidad']
            elif mov['tipo'] == 'salida':
                cant_saldo -= mov['cantidad']
                
        # Convertir la fila de la base de datos a un diccionario normal.
        material_info = dict(mat)
        # Añadir el stock calculado a este diccionario.
        material_info['stock'] = cant_saldo
        
        stock_materiales.append(material_info)
        
    cursor.close()
    conn.close()
    return render_template('consultor.html', materiales=stock_materiales)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=3000, debug=True)