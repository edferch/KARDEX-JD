"""Exportación del inventario y del kardex a Excel, y carga masiva de materiales
desde un archivo Excel."""
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from flask import Blueprint, request, redirect, url_for, render_template, flash, Response
from io import BytesIO

from ..db import get_db_connection
from ..inventarios import INVENTARIOS, obtener_inventario_actual
from ..logic import preparar_datos_kardex

excel_bp = Blueprint('excel_bp', __name__)


@excel_bp.route('/exportar_inventario')
def exportar_inventario():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT * FROM materiales WHERE inventario = ? ORDER BY nombre ASC', (obtener_inventario_actual(),))
    materiales = cursor.fetchall()
    cursor.close()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Estilos basados en la interfaz (CSS)
    fill_hdr_gris = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    font_hdr_gris = Font(color="475569", bold=True)
    alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    alignment_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    border_thin = Border(left=Side(style='thin', color='E2E8F0'),
                         right=Side(style='thin', color='E2E8F0'),
                         top=Side(style='thin', color='E2E8F0'),
                         bottom=Side(style='thin', color='E2E8F0'))

    # Las columnas 10 y 11 (Existencia Inicial, Costo Unitario) son numéricas; el resto es texto.
    # "Código" se agrega al final para no romper la carga masiva de plantillas ya existentes.
    columnas_numericas = {10, 11}
    headers = ['Nombre', 'Descripción', 'Grupo', 'No. Metrico', 'Origen', 'Fuente', 'Proveedor', 'Presentacion', 'Unidad', 'Existencia Inicial', 'Costo Unitario (Q)', 'Código']

    ws.append(headers)
    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = fill_hdr_gris
        cell.font = font_hdr_gris
        cell.alignment = alignment_right if col_num in columnas_numericas else alignment_left
        cell.border = border_thin

    for idx, mat in enumerate(materiales, 1):
        row = [mat['nombre'], mat['descripcion'], mat['tipo_material'], mat['numero_metrico'], mat['origen'], mat['fuente'], mat['empresa'], mat['presentacion'], mat['unidad'], mat['cantidad_inicial'], round(mat['precio_unitario'], 2), mat['codigo'] or '']
        ws.append(row)
        for col_num, cell in enumerate(ws[ws.max_row], 1):
            cell.alignment = alignment_right if col_num in columnas_numericas else alignment_left
            cell.border = border_thin
            if col_num in columnas_numericas:
                cell.number_format = '#,##0.00'
    # Ajustar ancho de las columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    output = BytesIO()
    wb.save(output)

    return Response(output.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', headers={'Content-Disposition': 'attachment; filename=Plantilla_Inventario.xlsx'})


@excel_bp.route('/exportar_kardex')
def exportar_kardex():
    mes_filtro = request.args.get('mes')
    if not mes_filtro:
        mes_filtro = datetime.now().strftime('%Y-%m')

    meses_es = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    if mes_filtro == 'todos':
        nombre_archivo = "Kardex General Completo.xlsx"
    else:
        try:
            anio, mes_num = mes_filtro.split('-')
            mes_nombre = meses_es[int(mes_num)]
            nombre_archivo = f"Kardex General Mes de {mes_nombre} de {anio}.xlsx"
        except Exception:
            nombre_archivo = "Kardex_General.xlsx"

    # Se calcula el kardex de los 3 inventarios (A, B, C), cada uno en su propia hoja,
    # más una hoja de resumen que totaliza compras, salidas y valor actual entre todos.
    conn = get_db_connection()
    cursor = conn.cursor()

    datos_por_inventario = {}
    for letra in INVENTARIOS:
        cursor.execute('SELECT * FROM materiales WHERE inventario = ? ORDER BY nombre ASC', (letra,))
        materiales = cursor.fetchall()

        movimientos_por_material = {}
        for mat in materiales:
            cursor.execute('SELECT * FROM movimientos WHERE material_id = ? ORDER BY fecha ASC, id ASC', (mat['id'],))
            movimientos_por_material[mat['id']] = cursor.fetchall()

        materiales_kardex, _, _, totales = preparar_datos_kardex(materiales, movimientos_por_material, mes_filtro)
        datos_por_inventario[letra] = {'materiales': materiales_kardex, 'totales': totales}

    cursor.close()
    conn.close()

    wb = openpyxl.Workbook()

    fill_gris = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_amarillo = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    fill_verde = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fill_azul = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    fill_naranja = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
    fill_total = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    bold_font = Font(bold=True)
    border = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'), top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))
    alignment_center = Alignment(horizontal='center', vertical='center')
    alignment_right = Alignment(horizontal='right', vertical='center')

    def _autoajustar_columnas(ws):
        for col in ws.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 40)

    # --- HOJA "Resumen General": total comprado, total salido y valor actual en materias primas ---
    ws_resumen = wb.active
    ws_resumen.title = "Resumen General"
    ws_resumen.append(['Inventario', 'Cant. Comprada', 'Total Comprado (Q)', 'Cant. Salida', 'Total Salido (Q)', 'Cant. Actual', 'Valor Actual en Materias Primas (Q)'])
    for cell in ws_resumen[1]:
        cell.fill = fill_gris
        cell.font = bold_font
        cell.border = border
        cell.alignment = alignment_center

    total_general = {'ing_cant': 0.0, 'ing_total': 0.0, 'sal_cant': 0.0, 'sal_total': 0.0, 'fin_cant': 0.0, 'fin_total': 0.0}
    for letra in INVENTARIOS:
        t = datos_por_inventario[letra]['totales']
        ws_resumen.append([f"Inventario {letra}", t['ing_cant'], t['ing_total'], t['sal_cant'], t['sal_total'], t['fin_cant'], t['fin_total']])
        for clave in total_general:
            total_general[clave] += t[clave]

    ws_resumen.append(['TOTAL GENERAL', total_general['ing_cant'], total_general['ing_total'], total_general['sal_cant'], total_general['sal_total'], total_general['fin_cant'], total_general['fin_total']])

    fila_total_resumen = ws_resumen.max_row
    for row in ws_resumen.iter_rows(min_row=2, max_row=ws_resumen.max_row):
        es_total = row[0].row == fila_total_resumen
        for col_num, cell in enumerate(row, 1):
            cell.border = border
            if col_num == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.number_format = '#,##0.00'
                cell.alignment = alignment_right
            if es_total:
                cell.font = bold_font
                cell.fill = fill_total

    _autoajustar_columnas(ws_resumen)

    # --- UNA HOJA POR INVENTARIO CON EL DETALLE DEL KARDEX ---
    headers = [
        'No.', 'Código', 'Nombre', 'Grupo',
        'Inicial Cant.', 'Inicial Costo', 'Inicial Total',
        'Entradas Cant.', 'Entradas Costo', 'Entradas Total',
        'Salidas Cant.', 'Salidas Costo', 'Salidas Total',
        'Actual Cant.', 'Actual Costo', 'Actual Total'
    ]

    for letra in INVENTARIOS:
        materiales_kardex = datos_por_inventario[letra]['materiales']
        totales = datos_por_inventario[letra]['totales']

        ws_kardex = wb.create_sheet(title=f"Kardex {letra}")
        ws_kardex.append(headers)
        for cell in ws_kardex[1]:
            cell.fill = fill_gris
            cell.font = bold_font
            cell.border = border
            cell.alignment = alignment_center

        for idx, mat in enumerate(materiales_kardex, 1):
            ws_kardex.append([
                idx, mat['codigo'] or '', mat['nombre'], mat['tipo_material'],
                mat['ini_cant'], mat['ini_costo'], mat['ini_total'],
                mat['ing_cant'], mat['ing_costo'], mat['ing_total'],
                mat['sal_cant'], mat['sal_costo'], mat['sal_total'],
                mat['fin_cant'], mat['fin_costo'], mat['fin_total']
            ])

        # Fila de totales generales de este inventario
        ws_kardex.append([
            '', '', 'TOTAL GENERAL', '',
            totales['ini_cant'], '', totales['ini_total'],
            totales['ing_cant'], '', totales['ing_total'],
            totales['sal_cant'], '', totales['sal_total'],
            totales['fin_cant'], '', totales['fin_total']
        ])
        fila_total = ws_kardex.max_row

        for row in ws_kardex.iter_rows(min_row=2, max_row=ws_kardex.max_row):
            es_total = row[0].row == fila_total
            for col_num, cell in enumerate(row, 1):
                cell.border = border
                if col_num in {5, 6, 7}:
                    cell.fill = fill_amarillo
                elif col_num in {8, 9, 10}:
                    cell.fill = fill_verde
                elif col_num in {11, 12, 13}:
                    cell.fill = fill_azul
                elif col_num in {14, 15, 16}:
                    cell.fill = fill_naranja

                if col_num == 1:
                    cell.alignment = alignment_center
                elif col_num >= 5:
                    cell.number_format = '#,##0.00'

                if es_total:
                    cell.font = bold_font
                    if col_num in (1, 2, 4, 6, 9, 12, 15):
                        cell.fill = fill_total

        ws_kardex.freeze_panes = 'A2'
        ws_kardex.auto_filter.ref = ws_kardex.dimensions
        _autoajustar_columnas(ws_kardex)

    output = BytesIO()
    wb.save(output)

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument/spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{nombre_archivo}"'}
    )


@excel_bp.route('/cargar_excel', methods=['GET', 'POST'])
def cargar_excel():
    if request.method == 'POST':
        if 'archivo_excel' not in request.files:
            flash('Error: No se encontró el archivo en la solicitud.', 'error')
            return redirect(request.url)

        file = request.files['archivo_excel']

        if file.filename == '':
            flash('Error: No se seleccionó ningún archivo.', 'error')
            return redirect(request.url)

        if file and file.filename.endswith('.xlsx'):
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                workbook = openpyxl.load_workbook(file)
                sheet = workbook.active

                rows_processed = 0
                rows_imported = 0
                rows_skipped = 0

                inventario_destino = request.form.get('inventario_destino')
                if inventario_destino not in INVENTARIOS:
                    inventario_destino = obtener_inventario_actual()

                sql_insert = '''
                    INSERT INTO materiales (nombre, descripcion, tipo_material, numero_metrico, origen, fuente, empresa, presentacion, unidad, cantidad_inicial, precio_unitario, drive_link, inventario, codigo)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                '''

                for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    rows_processed += 1
                    try:
                        if len(row) < 11:
                            rows_skipped += 1
                            continue

                        (nombre_raw, descripcion, tipo_material_raw, numero_metrico, origen,
                         fuente_raw, empresa_raw, presentacion, unidad, cantidad_inicial_raw, precio_unitario_raw) = row[:11]
                        # Columna 12 (Código) es opcional, para no romper plantillas viejas de 11 columnas.
                        codigo_raw = row[11] if len(row) > 11 else None

                        # --- VALIDACIÓN DE DATOS OBLIGATORIOS ---
                        if not all([nombre_raw, tipo_material_raw, fuente_raw, empresa_raw, cantidad_inicial_raw is not None, precio_unitario_raw is not None]):
                            rows_skipped += 1
                            continue

                        # --- LIMPIEZA Y CREACIÓN AUTOMÁTICA DE ENTIDADES ---
                        nombre = str(nombre_raw).strip()
                        tipo_material = str(tipo_material_raw).strip()
                        fuente = str(fuente_raw).strip()
                        empresa = str(empresa_raw).strip()

                        # En PostgreSQL, ON CONFLICT (nombre) DO NOTHING requiere que la columna tenga restricción UNIQUE
                        cursor.execute('INSERT INTO grupos (nombre) VALUES (?) ON CONFLICT (nombre) DO NOTHING', (tipo_material,))
                        cursor.execute('INSERT INTO fuentes (nombre) VALUES (?) ON CONFLICT (nombre) DO NOTHING', (fuente,))

                        cursor.execute('SELECT id FROM proveedores WHERE nombre = ?', (empresa,))
                        prov_exists = cursor.fetchone()
                        if not prov_exists:
                            cursor.execute('INSERT INTO proveedores (nombre, nit) VALUES (?, ?)', (empresa, ''))

                        cantidad_inicial = float(cantidad_inicial_raw)
                        precio_unitario = float(precio_unitario_raw)

                        codigo = str(codigo_raw).strip() if codigo_raw not in (None, '') else ''
                        values_to_insert = (nombre, descripcion, tipo_material, numero_metrico, origen, fuente, empresa, presentacion, unidad, cantidad_inicial, precio_unitario, '', inventario_destino, codigo)
                        cursor.execute(sql_insert, values_to_insert)
                        rows_imported += 1

                    except (ValueError, TypeError):
                        rows_skipped += 1
                        continue

                conn.commit()
                cursor.close()
                conn.close()

                flash_message = f"Éxito: Carga completada. Se importaron {rows_imported} materiales al Inventario {inventario_destino}."
                if rows_skipped > 0:
                    flash_message += f" Se omitieron {rows_skipped} filas por datos faltantes o formato incorrecto."
                flash(flash_message, "success")

            except Exception as e:
                flash(f"Error: Ocurrió un problema al procesar el archivo Excel: {e}", "error")
            return redirect(url_for('inventario_bp.inventario'))

    return render_template('carga_masiva.html')


@excel_bp.route('/exportar_actualizacion_materiales')
def exportar_actualizacion_materiales():
    """Plantilla para renombrar materiales existentes y asignarles código, sin
    tocar su descripción ni sus movimientos (se actualiza por ID, no se
    borra/recrea el material)."""
    inventario_sel = request.args.get('inventario')
    if inventario_sel not in INVENTARIOS:
        inventario_sel = obtener_inventario_actual()

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT id, codigo, nombre, descripcion FROM materiales WHERE inventario = ? ORDER BY nombre ASC', (inventario_sel,))
    materiales = cursor.fetchall()
    cursor.close()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Actualizar {inventario_sel}"

    fill_hdr = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_id = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    font_hdr = Font(color="475569", bold=True)
    font_gris = Font(color="94A3B8")
    alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border_thin = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
                         top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))

    headers = ['ID (no modificar)', 'Código', 'Nombre', 'Descripción actual (referencia, no se modifica)']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = fill_hdr
        cell.font = font_hdr
        cell.alignment = alignment_left
        cell.border = border_thin

    for mat in materiales:
        ws.append([mat['id'], mat['codigo'] or '', mat['nombre'], mat['descripcion'] or ''])
        fila = ws[ws.max_row]
        for cell in fila:
            cell.alignment = alignment_left
            cell.border = border_thin
        fila[0].fill = fill_id
        fila[0].font = font_gris
        fila[3].font = font_gris

    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    output = BytesIO()
    wb.save(output)

    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=Actualizar_Nombres_Codigos_Inventario_{inventario_sel}.xlsx'}
    )


@excel_bp.route('/actualizar_materiales_masivo', methods=['POST'])
def actualizar_materiales_masivo():
    """Carga masiva de SOLO nombre y código para materiales que ya existen
    (identificados por ID). No toca descripción, ni ningún otro campo, ni los
    movimientos del material, porque es un UPDATE sobre el mismo registro."""
    if 'archivo_excel' not in request.files:
        flash('Error: No se encontró el archivo en la solicitud.', 'error')
        return redirect(url_for('excel_bp.cargar_excel'))

    file = request.files['archivo_excel']
    if file.filename == '':
        flash('Error: No se seleccionó ningún archivo.', 'error')
        return redirect(url_for('excel_bp.cargar_excel'))

    if not file.filename.endswith('.xlsx'):
        flash('Error: El archivo debe ser un Excel (.xlsx).', 'error')
        return redirect(url_for('excel_bp.cargar_excel'))

    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active

        rows_processed = 0
        rows_updated = 0
        rows_skipped = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            rows_processed += 1
            try:
                if not row or len(row) < 3:
                    rows_skipped += 1
                    continue

                id_raw, codigo_raw, nombre_raw = row[0], row[1], row[2]

                if id_raw is None or not nombre_raw:
                    rows_skipped += 1
                    continue

                material_id = int(id_raw)
                nombre = str(nombre_raw).strip()
                codigo = str(codigo_raw).strip() if codigo_raw not in (None, '') else ''

                if not nombre:
                    rows_skipped += 1
                    continue

                cursor.execute('UPDATE materiales SET nombre = ?, codigo = ? WHERE id = ?', (nombre, codigo, material_id))
                if cursor.rowcount == 0:
                    rows_skipped += 1
                else:
                    rows_updated += 1

            except (ValueError, TypeError):
                rows_skipped += 1
                continue

        conn.commit()
        cursor.close()
        conn.close()

        flash_message = f"Éxito: Se actualizó el nombre y código de {rows_updated} materiales. Descripción y movimientos no se modificaron."
        if rows_skipped > 0:
            flash_message += f" Se omitieron {rows_skipped} filas por ID inválido o nombre vacío."
        flash(flash_message, "success")

    except Exception as e:
        flash(f"Error: Ocurrió un problema al procesar el archivo Excel: {e}", "error")

    return redirect(url_for('inventario_bp.inventario'))
